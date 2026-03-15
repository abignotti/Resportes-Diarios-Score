#!/usr/bin/env python3
"""
update_report.py
────────────────
Actualiza reportes de Score Energy (Cencosud + Walmart) y sincroniza a Supabase.

Uso:
  python update_report.py              # procesa Cencosud + Walmart
  python update_report.py --cencosud   # solo Cencosud
  python update_report.py --walmart    # solo Walmart
"""

import re
import time
import warnings
from datetime import date, timedelta
from pathlib import Path

import httpx
import openpyxl
import pandas as pd
from dotenv import load_dotenv
import os
import sys

warnings.filterwarnings("ignore")
load_dotenv()

# ── Rutas ──────────────────────────────────────────────────────────────────────
BASE_DIR = Path("/Users/agustin/Library/CloudStorage/OneDrive-uc.cl/Score/Dashboard Reportes Diarios")

# Cencosud
CENCOSUD_REPORT = BASE_DIR / "JUMBO_REPORTE.xlsx"
CENCOSUD_INPUTS = BASE_DIR / "Inputs_Cencosud"
CENCOSUD_CADENAS = {"JUMBO", "SANTA ISABEL", "SPID"}
CENCOSUD_SKUS = {2001113, 2034089, 2030990, 2030991,
                 2039041, 2048972, 1776033, 2001115, 2001114}

# Walmart
WALMART_REPORT = BASE_DIR / "WALMART_REPORTE.xlsx"
WALMART_INPUTS = BASE_DIR / "Inputs_Walmart"
WALMART_CADENAS = {"LIDER", "LIDER EXPRESS", "ACUENTA", "EKONO", "CENTRAL MAYORISTA"}

# Walmart: articulo_id canónico por sabor (Consumer ID representativo)
WALMART_SABOR_IDS = {
    "GORILLA":     92712338,
    "ORIGINAL":    220198871,
    "MANGO":       220198872,
    "ZERO":        220198873,
    "BUBBLE GUM":  224015661,
    "RADICAL":     226270474,
    "FRUIT PUNCH": 224015660,
    "CON GAS":     223158027,
    "SIN GAS":     223158028,
}
WALMART_SABOR_DESC = {
    "GORILLA":     "Beb Energetica Score Gorilla",
    "ORIGINAL":    "Beb Energetica Score Original",
    "MANGO":       "Beb Energetica Score Mango",
    "ZERO":        "Beb Energetica Score Zero",
    "BUBBLE GUM":  "Beb Energetica Score Bubblegum",
    "RADICAL":     "Beb Energetica Score Rad White",
    "FRUIT PUNCH": "Score Fruit Punch",
    "CON GAS":     "Score Agua Con Gas",
    "SIN GAS":     "Score Agua Sin Gas",
}

# ── Supabase ───────────────────────────────────────────────────────────────────
SUPABASE_URL         = os.getenv("SUPABASE_URL", "")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")
SUPABASE_ENABLED     = bool(SUPABASE_URL and SUPABASE_SERVICE_KEY)


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS COMUNES
# ══════════════════════════════════════════════════════════════════════════════

def latest(directory: Path, prefix: str) -> Path:
    files = sorted(p for p in directory.iterdir() if p.name.startswith(prefix))
    if not files:
        raise FileNotFoundError(f"No se encontró '{prefix}*' en {directory}")
    return files[-1]


def drop_blank_cols(df: pd.DataFrame) -> pd.DataFrame:
    return df.drop(
        columns=[c for c in df.columns if str(c).strip() in ("", " ", "nan")],
        errors="ignore",
    )


def clean_val(v):
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime() if pd.notna(v) else None
    if hasattr(v, "item"):
        return v.item()
    return v


def df_to_records(df: pd.DataFrame) -> list[dict]:
    records = []
    for row in df.to_dict("records"):
        records.append({k: clean_val(v) for k, v in row.items()})
    return records


def sb_upsert(table: str, records: list[dict], on_conflict: str) -> None:
    if not SUPABASE_ENABLED or not records:
        return
    headers = {
        "apikey":        SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type":  "application/json",
        "Prefer":        "resolution=merge-duplicates,return=minimal",
    }
    batch_size = 500
    total = len(records)
    for i in range(0, total, batch_size):
        batch = records[i : i + batch_size]
        ok = False
        for attempt in range(3):
            try:
                r = httpx.post(
                    f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={on_conflict}",
                    headers=headers, json=batch, timeout=60,
                )
                if r.status_code not in (200, 201):
                    print(f"  ⚠️  Supabase error {table} lote {i//batch_size+1}: {r.text[:200]}")
                else:
                    end = min(i + batch_size, total)
                    print(f"  ✓ Supabase {table}: filas {i+1}–{end} de {total}")
                ok = True
                break
            except (httpx.ConnectError, httpx.ReadError) as e:
                if attempt < 2:
                    wait = (attempt + 1) * 2
                    print(f"  ⏳ Conexión perdida lote {i//batch_size+1}, reintentando en {wait}s... ({e})")
                    time.sleep(wait)
                else:
                    print(f"  ⚠️  Falló lote {i//batch_size+1} tras 3 intentos: {e}")
        if ok and total > batch_size:
            time.sleep(0.3)


def calc_daily_sales(ventas: pd.DataFrame, today: date) -> pd.DataFrame:
    """Calcula D0–D13 pivotando ventas por fecha."""
    pivot = ventas.pivot_table(
        index=["Local ID", "Artículo ID"],
        columns="Fecha",
        values="Venta Unidades",
        aggfunc="sum",
        fill_value=0,
    )
    result = pd.DataFrame(index=pivot.index)
    for i in range(14):
        target = pd.Timestamp(today) - timedelta(days=i + 1)
        result[f"D{i}"] = pivot[target] if target in pivot.columns else 0
    return result.reset_index()


def stock_to_sb(stock: pd.DataFrame) -> list[dict]:
    """Normaliza stock DataFrame a formato Supabase (funciona para ambas cadenas)."""
    cols = {
        "Día":                   "dia",
        "Cadena":                "cadena",
        "Local ID":              "local_id",
        "Local DESC":            "local_desc",
        "Artículo ID":           "articulo_id",
        "Artículo DESC":         "articulo_desc",
        "Stock Disponibilidad":  "stock_disponibilidad",
        "Stock LU":              "stock_lu",
        "Stock Tránsito":        "stock_transito",
        "Forecast":              "forecast",
        "Cubicaje":              "cubicaje",
        "Min Exhibición":        "min_exhibicion",
        "D0":"d0","D1":"d1","D2":"d2","D3":"d3","D4":"d4","D5":"d5","D6":"d6",
        "D7":"d7","D8":"d8","D9":"d9","D10":"d10","D11":"d11","D12":"d12","D13":"d13",
        "L7D":                   "l7d",
        "L14D":                  "l14d",
        "Alerta Agencia":        "alerta_agencia",
        "ALERTA REPOSICIÓN":     "alerta_reposicion",
        "Región":                "region",
        "SUPERVISOR":            "supervisor",
        "ATENCION":              "atencion",
        "INCENTIVO":             "incentivo",
    }
    df = stock[[c for c in cols if c in stock.columns]].rename(columns=cols)
    if "dia" in df.columns:
        df["dia"] = pd.to_datetime(df["dia"]).dt.strftime("%Y-%m-%d")
    return df_to_records(df)


def ventas_to_sb(ventas: pd.DataFrame) -> list[dict]:
    """Normaliza ventas DataFrame a formato Supabase."""
    cols = {
        "Fecha":              "fecha",
        "Cadena":             "cadena",
        "Local ID":           "local_id",
        "Local DESC":         "local_desc",
        "Artículo ID":        "articulo_id",
        "Artículo DESC":      "articulo_desc",
        "Venta Unidades":     "venta_unidades",
        "Venta Neta ($)":     "venta_neta",
        "Costo Venta ($)":    "costo_venta",
        "Contribución ($)":   "contribucion",
        "Región":             "region",
    }
    df = ventas[[c for c in cols if c in ventas.columns]].rename(columns=cols)
    if "fecha" in df.columns:
        df["fecha"] = df["fecha"].dt.strftime("%Y-%m-%d")
    # Deduplicar
    num_cols = ["venta_unidades", "venta_neta", "costo_venta", "contribucion"]
    key_cols = ["fecha", "local_id", "articulo_id"]
    meta_cols = [c for c in df.columns if c not in num_cols + key_cols]
    agg = {c: "sum" for c in num_cols if c in df.columns}
    agg.update({c: "first" for c in meta_cols if c in df.columns})
    df = df.groupby(key_cols, as_index=False).agg(agg)
    return df_to_records(df)


def write_sheet(wb: openpyxl.Workbook, sheet_name: str, df: pd.DataFrame, header_row: int = 2) -> None:
    ws = wb[sheet_name]
    header_vals = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)
    df_cols = set(df.columns)
    for _, row in df.iterrows():
        row_data = [
            clean_val(row[str(h).strip()]) if h and str(h).strip() in df_cols else None
            for h in header_vals
        ]
        ws.append(row_data)
    print(f"  ✓ Excel {sheet_name}: {len(df):,} filas")


# ══════════════════════════════════════════════════════════════════════════════
#  CENCOSUD
# ══════════════════════════════════════════════════════════════════════════════

def read_cencosud_ventas() -> pd.DataFrame:
    path = latest(CENCOSUD_INPUTS, "Ventas por Art")
    print(f"  📄 {path.name}")
    df = pd.read_excel(path)
    df = drop_blank_cols(df)
    df = df[df["Fecha"] != "Total"].copy()
    df["Artículo ID"] = pd.to_numeric(df["Artículo ID"], errors="coerce").astype("Int64")
    df = df[df["Cadena"].isin(CENCOSUD_CADENAS) & df["Artículo ID"].isin(CENCOSUD_SKUS)].copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.rename(columns={"Local": "Local DESC", "Artículo": "Artículo DESC"}, errors="ignore")
    df["SKU"] = df["Artículo ID"]
    return df


def read_cencosud_stock() -> pd.DataFrame:
    path = latest(CENCOSUD_INPUTS, "Maestra - Stock Detalle")
    print(f"  📄 {path.name}")
    df = pd.read_excel(path)
    df = drop_blank_cols(df)
    df = df[df["Día"] != "Total"].copy()
    df["Artículo ID"] = pd.to_numeric(df["Artículo ID"], errors="coerce").astype("Int64")
    df = df[df["Cadena"].isin(CENCOSUD_CADENAS) & df["Artículo ID"].isin(CENCOSUD_SKUS)].copy()
    df["Día"] = pd.to_datetime(df["Día"], errors="coerce")
    return df


def read_cencosud_maestro() -> pd.DataFrame:
    return pd.read_excel(CENCOSUD_REPORT, sheet_name="MAESTRO")


def add_cencosud_calculated_cols(stock, ventas, maestro, today):
    dsales = calc_daily_sales(ventas, today)
    df = stock.merge(dsales, on=["Local ID", "Artículo ID"], how="left")
    d_cols = [f"D{i}" for i in range(14)]
    df[d_cols] = df[d_cols].fillna(0).astype(int)
    df["L7D"]  = df[[f"D{i}" for i in range(7)]].sum(axis=1)
    df["L14D"] = df[d_cols].sum(axis=1)

    m = maestro.copy()
    m["Store Nbr"] = m["Store Nbr"].astype(str).str.strip()
    lid = df["Local ID"].astype(str).str.strip()

    sup_map  = m.dropna(subset=["Encargado"]).set_index("Store Nbr")["Encargado"].to_dict()
    aten_map = m.dropna(subset=["Dia"]).set_index("Store Nbr")["Dia"].to_dict()
    inc_map  = m.dropna(subset=["INCENTIVO"]).set_index("SALA")["INCENTIVO"].to_dict()

    df["SUPERVISOR"] = lid.map(sup_map).fillna("").replace(0, "")
    df["ATENCION"]   = lid.map(aten_map).fillna("").replace(0, "")
    df["INCENTIVO"]  = lid.map(inc_map).fillna("").replace(0, "")

    region_map = (
        ventas.dropna(subset=["Región"])
        .drop_duplicates("Local ID")
        .set_index("Local ID")["Región"]
        .to_dict()
    )
    df["Región"] = df["Local ID"].map(region_map).fillna("")

    sd = df["Stock Disponibilidad"].fillna(0)
    df["Alerta Agencia"]    = ((sd > 24)  & (df["D0"] + df["D1"] == 0)).astype(int)
    df["ALERTA REPOSICIÓN"] = ((sd >= 24) & (df["D0"] == 0)).astype(int)
    return df


def cencosud_maestro_to_sb(maestro):
    cencosud = maestro[maestro["Cadena"].str.upper().isin(CENCOSUD_CADENAS)].copy()
    keep = {
        "Store Nbr": "local_id", "Cadena": "cadena", "Store Name": "store_name",
        "COMUNA": "comuna", "Encargado": "encargado", "Incentivo": "incentivo",
        "SUPERVISOR": "supervisor",
    }
    available = {k: v for k, v in keep.items() if k in cencosud.columns}
    df = cencosud[list(available.keys())].rename(columns=available)
    df["sala"] = df["local_id"]
    df = df.dropna(subset=["local_id"])
    return df_to_records(df)


def process_cencosud(today):
    sep = "─" * 50
    print(f"\n{sep}")
    print("  🟢 CENCOSUD (Jumbo / Santa Isabel / SPID)")
    print(f"{sep}")

    print("\n📂 Leyendo inputs Cencosud...")
    ventas  = read_cencosud_ventas()
    stock   = read_cencosud_stock()
    maestro = read_cencosud_maestro()

    print(f"     Ventas : {len(ventas):,} filas | "
          f"{ventas['Fecha'].min():%d/%m} → {ventas['Fecha'].max():%d/%m/%Y}")
    print(f"     Stock  : {len(stock):,} filas")

    print("\n🧮 Calculando columnas derivadas...")
    stock_final = add_cencosud_calculated_cols(stock, ventas, maestro, today)

    print("\n💾 Actualizando JUMBO_REPORTE.xlsx...")
    wb = openpyxl.load_workbook(CENCOSUD_REPORT)
    write_sheet(wb, "Venta", ventas)
    write_sheet(wb, "Stock", stock_final)
    wb.save(CENCOSUD_REPORT)

    if SUPABASE_ENABLED:
        print("\n☁️  Sincronizando Cencosud a Supabase...")
        sb_upsert("maestro", cencosud_maestro_to_sb(maestro), "local_id")
        sb_upsert("ventas",  ventas_to_sb(ventas),             "fecha,local_id,articulo_id")
        sb_upsert("stock",   stock_to_sb(stock_final),         "dia,local_id,articulo_id")

    alerts_ag   = stock_final["Alerta Agencia"].sum()
    alerts_repo = stock_final["ALERTA REPOSICIÓN"].sum()
    print(f"\n  ✅ Cencosud OK — {len(stock_final):,} filas | "
          f"Alertas: {alerts_ag} ag. / {alerts_repo} repo.")


# ══════════════════════════════════════════════════════════════════════════════
#  WALMART
# ══════════════════════════════════════════════════════════════════════════════

def build_walmart_maps():
    """Lee MAESTRO Walmart y construye los mapeos necesarios."""
    maestro = pd.read_excel(WALMART_REPORT, sheet_name="MAESTRO", header=1)
    maestro = drop_blank_cols(maestro)

    # Item Nbr → sabor
    item_to_sabor = {}
    for _, row in maestro.iterrows():
        item = row.get("Item Nbr")
        sabor = row.get("SABOR")
        if pd.notna(item) and pd.notna(sabor):
            item_to_sabor[int(item)] = str(sabor).strip().upper()

    # Store Nbr → cadena (normalizada)
    store_to_cadena = {}
    store_to_name = {}
    for _, row in maestro.iterrows():
        store = row.get("Store Nbr")
        cadena = row.get("Cadena")
        name = row.get("Store Name")
        if pd.notna(store) and pd.notna(cadena):
            c = str(cadena).strip().upper()
            if c == "LIDER EXPRES":
                c = "LIDER EXPRESS"
            store_to_cadena[int(store)] = c
            if pd.notna(name):
                store_to_name[int(store)] = str(name).strip()

    # Maestro compatible para lookups de supervisor
    m_compat = maestro.copy()
    if "SUPERVISOR" in m_compat.columns:
        m_compat["Encargado"] = m_compat["SUPERVISOR"]
    m_compat["Store Nbr"] = m_compat["Store Nbr"].astype(str).str.strip()

    return item_to_sabor, store_to_cadena, store_to_name, m_compat


def parse_walmart_date(path: Path) -> date:
    """Extrae la fecha del stock desde los metadatos del archivo Walmart."""
    meta = pd.read_excel(path, header=None, nrows=15)
    for _, row in meta.iterrows():
        for val in row:
            if pd.notna(val) and "Pos Date" in str(val):
                m = re.search(r"(\d{2})-(\d{2})-(\d{4})", str(val))
                if m:
                    return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
    # Fallback: parsear desde nombre (DD-MM-YYYY)
    m = re.search(r"(\d{2})-(\d{2})-(\d{4})", path.name)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return date.today() - timedelta(days=1)


def read_walmart_sellout(item_to_sabor, store_to_cadena, store_to_name):
    """Lee Sell Out de Walmart (header en fila 17, 0-indexed)."""
    path = latest(WALMART_INPUTS, "Sell Out")
    print(f"  📄 {path.name}")

    df = pd.read_excel(path, header=17)
    df = drop_blank_cols(df)

    df["Daily"] = pd.to_datetime(df["Daily"], errors="coerce")
    df = df.dropna(subset=["Daily"])

    # Solo últimos 14 días
    max_date = df["Daily"].max()
    cutoff = max_date - timedelta(days=13)
    df = df[df["Daily"] >= cutoff].copy()

    # Item Nbr → sabor → articulo_id canónico
    df["Item Nbr"] = pd.to_numeric(df["Item Nbr"], errors="coerce")
    df["sabor"] = df["Item Nbr"].map(item_to_sabor)
    df = df.dropna(subset=["sabor"])
    df["articulo_id"] = df["sabor"].map(WALMART_SABOR_IDS)

    # Store Nbr → cadena
    df["Store Nbr"] = pd.to_numeric(df["Store Nbr"], errors="coerce").astype("Int64")
    df["cadena"] = df["Store Nbr"].map(store_to_cadena)
    df = df[df["cadena"].isin(WALMART_CADENAS)].copy()

    # Agregar por (Store, articulo_id, fecha)
    agg = df.groupby(["Store Nbr", "articulo_id", "Daily"]).agg({
        "POS Qty": "sum",
        "POS Sales": "sum",
        "Store Name": "first",
        "Region": "first",
        "cadena": "first",
        "sabor": "first",
    }).reset_index()

    result = pd.DataFrame({
        "Fecha":           agg["Daily"],
        "Cadena":          agg["cadena"],
        "Local ID":        agg["Store Nbr"].astype(str),
        "Local DESC":      agg["Store Nbr"].map(store_to_name).fillna(agg["Store Name"]),
        "Artículo ID":     agg["articulo_id"],
        "Artículo DESC":   agg["sabor"].map(WALMART_SABOR_DESC),
        "Venta Unidades":  agg["POS Qty"],
        "Venta Neta ($)":  agg["POS Sales"],
        "Región":          agg["Region"],
        "SKU":             agg["articulo_id"],
    })
    return result


def read_walmart_stock(item_to_sabor, store_to_cadena, store_to_name):
    """Lee Stock de Walmart (header en fila 19, 0-indexed)."""
    path = latest(WALMART_INPUTS, "Stock")
    print(f"  📄 {path.name}")

    stock_date = parse_walmart_date(path)
    print(f"     Fecha stock: {stock_date}")

    df = pd.read_excel(path, header=19)
    df = drop_blank_cols(df)

    df["Item Nbr"] = pd.to_numeric(df["Item Nbr"], errors="coerce")
    df["sabor"] = df["Item Nbr"].map(item_to_sabor)
    df = df.dropna(subset=["sabor"])
    df["articulo_id"] = df["sabor"].map(WALMART_SABOR_IDS)

    df["Store Nbr"] = pd.to_numeric(df["Store Nbr"], errors="coerce").astype("Int64")
    df["cadena"] = df["Store Nbr"].map(store_to_cadena)
    df = df[df["cadena"].isin(WALMART_CADENAS)].copy()

    agg = df.groupby(["Store Nbr", "articulo_id"]).agg({
        "Curr Str On Hand Qty":    "sum",
        "Curr Str In Transit Qty": "sum",
        "Curr Str In Whse Qty":    "sum",
        "Max Shelf Qty":           "max",
        "Store Name":              "first",
        "Region":                  "first",
        "cadena":                  "first",
        "sabor":                   "first",
    }).reset_index()

    result = pd.DataFrame({
        "Día":                  pd.Timestamp(stock_date),
        "Cadena":               agg["cadena"],
        "Local ID":             agg["Store Nbr"].astype(str),
        "Local DESC":           agg["Store Nbr"].map(store_to_name).fillna(agg["Store Name"]),
        "Artículo ID":          agg["articulo_id"],
        "Artículo DESC":        agg["sabor"].map(WALMART_SABOR_DESC),
        "Stock Disponibilidad": agg["Curr Str On Hand Qty"],
        "Stock Tránsito":       agg["Curr Str In Transit Qty"],
        "Min Exhibición":       agg["Max Shelf Qty"],
    })
    return result, stock_date


def add_walmart_calculated_cols(stock, ventas, wm_maestro, today):
    """Calcula D0–D13, L7D, L14D, alertas y lookups para Walmart."""
    dsales = calc_daily_sales(ventas, today)
    df = stock.merge(dsales, on=["Local ID", "Artículo ID"], how="left")
    d_cols = [f"D{i}" for i in range(14)]
    df[d_cols] = df[d_cols].fillna(0).astype(int)
    df["L7D"]  = df[[f"D{i}" for i in range(7)]].sum(axis=1)
    df["L14D"] = df[d_cols].sum(axis=1)

    # Lookups
    m = wm_maestro.copy()
    lid = df["Local ID"].astype(str).str.strip()
    if "Encargado" in m.columns:
        sup_map = m.dropna(subset=["Encargado"]).drop_duplicates("Store Nbr").set_index("Store Nbr")["Encargado"].to_dict()
        df["SUPERVISOR"] = lid.map(sup_map).fillna("")
    else:
        df["SUPERVISOR"] = ""

    if "Región" in ventas.columns:
        region_map = (
            ventas.dropna(subset=["Región"])
            .drop_duplicates("Local ID")
            .set_index("Local ID")["Región"]
            .to_dict()
        )
        df["Región"] = df["Local ID"].map(region_map).fillna("")
    else:
        df["Región"] = ""

    df["ATENCION"] = ""
    df["INCENTIVO"] = ""

    sd = df["Stock Disponibilidad"].fillna(0)
    df["Alerta Agencia"]    = ((sd > 24)  & (df["D0"] + df["D1"] == 0)).astype(int)
    df["ALERTA REPOSICIÓN"] = ((sd >= 24) & (df["D0"] == 0)).astype(int)
    return df


def walmart_maestro_to_sb(wm_maestro, store_to_cadena):
    """Prepara MAESTRO Walmart para Supabase."""
    rows = []
    seen = set()
    for _, r in wm_maestro.iterrows():
        store = r.get("Store Nbr")
        if pd.isna(store) or str(store).strip() in seen:
            continue
        store_str = str(store).strip()
        seen.add(store_str)
        store_int = int(float(store_str)) if store_str.replace(".", "").replace("-", "").isdigit() else None
        cadena = store_to_cadena.get(store_int, "") if store_int else ""
        if cadena not in WALMART_CADENAS:
            continue
        rows.append({
            "local_id":   store_str,
            "cadena":     cadena,
            "store_name": clean_val(r.get("Store Name")),
            "comuna":     clean_val(r.get("COMUNA")),
            "encargado":  clean_val(r.get("Encargado")),
            "supervisor": clean_val(r.get("SUPERVISOR", r.get("Encargado"))),
            "sala":       store_str,
        })
    return rows


def process_walmart(today):
    sep = "─" * 50
    print(f"\n{sep}")
    print("  🔵 WALMART (Lider / Lider Express / Acuenta / Ekono / C.Mayorista)")
    print(f"{sep}")

    print("\n📂 Construyendo mapeos desde MAESTRO Walmart...")
    item_to_sabor, store_to_cadena, store_to_name, wm_maestro = build_walmart_maps()
    print(f"     {len(item_to_sabor)} Item Nbr → {len(set(item_to_sabor.values()))} sabores")
    print(f"     {len(store_to_cadena)} tiendas mapeadas")

    print("\n📂 Leyendo inputs Walmart...")
    ventas = read_walmart_sellout(item_to_sabor, store_to_cadena, store_to_name)
    stock, stock_date = read_walmart_stock(item_to_sabor, store_to_cadena, store_to_name)

    # "today" para Walmart = stock_date + 1 (D0 = stock_date)
    wm_today = stock_date + timedelta(days=1)

    print(f"     Ventas : {len(ventas):,} filas | "
          f"{ventas['Fecha'].min():%d/%m} → {ventas['Fecha'].max():%d/%m/%Y}")
    print(f"     Stock  : {len(stock):,} filas (fecha: {stock_date})")

    print("\n🧮 Calculando columnas derivadas Walmart...")
    stock_final = add_walmart_calculated_cols(stock, ventas, wm_maestro, wm_today)

    if WALMART_REPORT.exists():
        print("\n💾 Actualizando WALMART_REPORTE.xlsx...")
        try:
            wb = openpyxl.load_workbook(WALMART_REPORT)
            if "SELLOUT" in wb.sheetnames:
                write_sheet(wb, "SELLOUT", ventas)
            if "STOCK" in wb.sheetnames:
                write_sheet(wb, "STOCK", stock_final)
            wb.save(WALMART_REPORT)
        except Exception as e:
            print(f"  ⚠️  No se pudo actualizar Excel: {e}")

    if SUPABASE_ENABLED:
        print("\n☁️  Sincronizando Walmart a Supabase...")
        sb_upsert("maestro", walmart_maestro_to_sb(wm_maestro, store_to_cadena), "local_id")
        sb_upsert("ventas",  ventas_to_sb(ventas),       "fecha,local_id,articulo_id")
        sb_upsert("stock",   stock_to_sb(stock_final),   "dia,local_id,articulo_id")

    alerts_ag   = stock_final["Alerta Agencia"].sum()
    alerts_repo = stock_final["ALERTA REPOSICIÓN"].sum()
    print(f"\n  ✅ Walmart OK — {len(stock_final):,} filas | "
          f"Alertas: {alerts_ag} ag. / {alerts_repo} repo.")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    today = date.today()
    args = set(sys.argv[1:])
    run_cencosud = "--cencosud" in args or not args
    run_walmart  = "--walmart"  in args or not args

    sep = "═" * 54
    print(f"\n{sep}")
    print(f"  Score Energy — Actualización de Reportes")
    print(f"  Fecha: {today:%d/%m/%Y}")
    chains = []
    if run_cencosud:
        chains.append("Cencosud")
    if run_walmart:
        chains.append("Walmart")
    print(f"  Cadenas: {' + '.join(chains)}")
    print(f"  Supabase: {'✅ activo' if SUPABASE_ENABLED else '⚠️  desactivado'}")
    print(f"{sep}")

    if run_cencosud:
        try:
            process_cencosud(today)
        except FileNotFoundError as e:
            print(f"\n  ⚠️  Cencosud omitido: {e}")

    if run_walmart:
        try:
            process_walmart(today)
        except FileNotFoundError as e:
            print(f"\n  ⚠️  Walmart omitido: {e}")

    print(f"\n{sep}")
    print(f"  🏁 Proceso finalizado")
    print(f"{sep}\n")


if __name__ == "__main__":
    main()
